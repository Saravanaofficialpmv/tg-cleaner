"""
Dashboard routes: fetch chats, manage keep-list, export.
"""

import logging
import csv
import io
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete

from app.database import get_db
from app.services import telegram as tg_service
from app.services import session as sess_service
from app.models.models import ChatRecord

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/dashboard", tags=["dashboard"])


# ── Helpers ───────────────────────────────────────────────────────────────────

async def _require_session(session_id: str, db: AsyncSession):
    """Retrieve session or raise 401."""
    session = await sess_service.get_user_session(db=db, session_id=session_id)
    if not session:
        raise HTTPException(status_code=401, detail="Session not found or expired.")
    return session


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/chats")
async def get_chats(
    session_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Fetch all groups/channels/supergroups from Telegram and cache them.
    Returns the list immediately from DB if already fetched; use ?refresh=true to re-fetch.
    """
    session = await _require_session(session_id, db)

    # Check DB cache first
    existing = await db.execute(
        select(ChatRecord).where(ChatRecord.session_id == session_id)
    )
    cached = existing.scalars().all()
    if cached:
        return {"chats": [_chat_to_dict(c) for c in cached], "cached": True}

    # Fetch from Telegram
    try:
        try:
            client = await tg_service.get_client_for_session(session_id, db)
        except Exception as e:
            logger.error(f"Failed to get client: {e}")
            raise HTTPException(
                status_code=400,
                detail=f"Telegram client not connected. Please log in again. Error: {str(e)}"
            )

        chats_data = await tg_service.fetch_dialogs(client)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching dialogs: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch chats: {str(e)}")

    # Persist to DB
    records = []
    for chat in chats_data:
        last_activity = None
        if chat.get("last_activity"):
            try:
                last_activity = datetime.fromisoformat(chat["last_activity"])
            except Exception:
                pass

        record = ChatRecord(
            session_id=session_id,
            chat_id=chat["chat_id"],
            name=chat["name"],
            chat_type=chat["chat_type"],
            username=chat.get("username"),
            member_count=chat.get("member_count"),
            unread_count=chat.get("unread_count", 0),
            is_muted=chat.get("is_muted", False),
            is_archived=chat.get("is_archived", False),
            last_activity=last_activity,
            is_kept=False,
        )
        db.add(record)
        records.append(record)

    await db.flush()
    return {"chats": [_chat_to_dict(r) for r in records], "cached": False}


@router.post("/refresh")
async def refresh_chats(
    body: dict,
    db: AsyncSession = Depends(get_db),
):
    """Delete cached records and re-fetch from Telegram."""
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required.")

    await _require_session(session_id, db)

    # Delete existing cache
    await db.execute(
        delete(ChatRecord).where(ChatRecord.session_id == session_id)
    )
    await db.flush()

    return {"message": "Cache cleared. Fetch /dashboard/chats to reload."}


class UpdateKeepRequest(BaseModel):
    session_id: str
    chat_ids: List[int]
    keep: bool


@router.post("/keep")
async def update_keep_list(
    body: UpdateKeepRequest,
    db: AsyncSession = Depends(get_db),
):
    """Toggle keep status for a list of chat IDs."""
    await _require_session(body.session_id, db)

    result = await db.execute(
        select(ChatRecord).where(
            ChatRecord.session_id == body.session_id,
            ChatRecord.chat_id.in_(body.chat_ids),
        )
    )
    records = result.scalars().all()
    for record in records:
        record.is_kept = body.keep

    await db.flush()
    return {"updated": len(records), "keep": body.keep}


@router.get("/export")
async def export_keep_list(
    session_id: str = Query(...),
    format: str = Query("csv"),
    db: AsyncSession = Depends(get_db),
):
    """Export keep list as CSV or Excel."""
    await _require_session(session_id, db)

    result = await db.execute(
        select(ChatRecord).where(
            ChatRecord.session_id == session_id,
            ChatRecord.is_kept == True,
        )
    )
    records = result.scalars().all()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Name", "Type", "Members", "Unread", "Muted", "Last Activity"])
        for r in records:
            writer.writerow([
                r.name, r.chat_type, r.member_count or "",
                r.unread_count, r.is_muted,
                r.last_activity.isoformat() if r.last_activity else "",
            ])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=keep_list.csv"},
        )

    # Excel export
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keep List"
    headers = ["Name", "Type", "Members", "Unread", "Muted", "Last Activity"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="1a1f35", end_color="1a1f35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        ws.append([
            r.name, r.chat_type, r.member_count or 0,
            r.unread_count, "Yes" if r.is_muted else "No",
            r.last_activity.isoformat() if r.last_activity else "",
        ])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=keep_list.xlsx"},
    )


def _chat_to_dict(record: ChatRecord) -> dict:
    return {
        "id": record.id,
        "chat_id": record.chat_id,
        "name": record.name,
        "chat_type": record.chat_type,
        "username": record.username,
        "member_count": record.member_count,
        "unread_count": record.unread_count,
        "is_muted": record.is_muted,
        "is_archived": record.is_archived,
        "last_activity": record.last_activity.isoformat() if record.last_activity else None,
        "is_kept": record.is_kept,
    }
